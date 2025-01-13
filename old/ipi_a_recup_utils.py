from bs4 import BeautifulSoup
import openpyxl
import pyautogui
import tkinter as tk
from tkinter import simpledialog
import time
import os
import pygetwindow as gw

def get_user_input():
    root = tk.Tk()
    root.withdraw()  # Esconder a janela principal
    
    # Solicitar o nome da empresa
    company_name = simpledialog.askstring(title="Nome da Empresa", prompt="Digite o nome da empresa:")

    # Solicitar o código da empresa
    company_code = simpledialog.askstring(title="Código da Empresa", prompt="Digite o código da empresa:")

    # Solicitar o mês e o ano
    month_year = simpledialog.askstring(title="Mês e Ano", prompt="Digite o mês e o ano (MMYYYY):")

    return company_code, month_year, company_name

def find_icms(html_icms_recup):
    
    company_code, month_year, company_name = get_user_input()
    html_path_icms_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração DimeSC Sequência 22 - Ordem 1.htm"

    try:
        with open(html_path_icms_recup, 'r', encoding='utf-8') as file:
            content = file.read()

        # Skip analysis if the file has less than 100 lines
        if len(content.splitlines()) < 100:
            print("HTML file has less than 100 lines, skipping analysis.")
        else:
            # Cria um objeto BeautifulSoup
            soup = BeautifulSoup(content, 'html.parser')

            # Busca a linha que começa com <td colspan="2" class="s9">190</td>
            result = soup.find('td', {'colspan': '2', 'class': 's9'}, string='190')

            # Exibe a linha inteira
            if result:
                row = result.find_parent('tr')
                observations = row.find_all('td')
                if len(observations) > 1:
                    ICMS_recup = observations[2].text
                    ICMS_recup = ICMS_recup.replace('.', '').replace(',', '.')
                    ICMS_recup = float(ICMS_recup)
                    print(f"ICMS_recup: {ICMS_recup:.2f}".replace('.', ','))
    except FileNotFoundError:
        print(f"Erro: O arquivo {html_path_icms_recup} não existe. Continuando a execução...")

    # Abrir a planilha e procurar pelos números na coluna A
    excel_path = f'C:\\projeto\\planilhas\\balancete\\CONCILIACAO_{company_code}_{month_year}.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
        
    numeros_procurados = [41, 42]

def extract_saldo_credor(html_path_ipi_a_recup):
    try:
        with open(html_path_ipi_a_recup, 'r', encoding='utf-8') as file:
            html_content = file.read()
            
        soup = BeautifulSoup(html_content, 'html.parser')
        saldo_credor_element = soup.find('td', string="SALDO CREDOR PERIODO SEGUINTE")

        if saldo_credor_element:
            ipi_a_recup = saldo_credor_element.find_next_sibling('td').get_text(strip=True)
            ipi_a_recup = float(ipi_a_recup.replace('.', '').replace(',', '.'))
            return ipi_a_recup
        else:
            print("Variável 'SALDO CREDOR PERIODO SEGUINTE' não encontrada.")
            return None
    except FileNotFoundError:
        print(f"Erro: O arquivo {html_path_ipi_a_recup} não existe. Continuando a execução...")
        return None

def extract_saldo_credor_cofins_recup(html_path_cofins_recup):
    try:
        with open(html_path_cofins_recup, 'r', encoding='utf-8') as file:
            html_content = file.read()
            
        soup = BeautifulSoup(html_content, 'html.parser')
        saldo_credor_element = soup.find('td', string="SALDO CREDOR PERIODO SEGUINTE")

        if saldo_credor_element:
            cofins_recup = saldo_credor_element.find_next_sibling('td').get_text(strip=True)
            cofins_recup = float(cofins_recup.replace('.', '').replace(',', '.'))
            return cofins_recup
        else:
            print("Variável 'SALDO CREDOR PERIODO SEGUINTE' não encontrada.")
            return None
    except FileNotFoundError:
        print(f"Erro: O arquivo {html_path_cofins_recup} não existe. Continuando a execução...")
        return None

def compare_and_write_to_excel(ws, numeros_procurados, ICMS_recup, ipi_a_recup, cofins_recup):
    for row in ws.iter_rows(min_row=2):
        cell_a = row[0].value  # Coluna A (índice 0)
        if cell_a in numeros_procurados:
            valor_coluna_h = row[7].value  # Coluna H (índice 7)
            print(f"Número {cell_a} encontrado: Valor na coluna H = {valor_coluna_h}")
            
            # Comparar os valores e escrever "OK" ou "Verificar" na coluna I
            try:
                if cell_a == 41:
                    try:
                        if abs(valor_coluna_h - ICMS_recup) <= 0.10:
                            row[8].value = "OK"
                        else:
                            row[8].value = "Verificar"
                    except NameError:
                        print("Erro: ICMS_recup não definido. Continuando a execução...")
                        row[8].value = "Verificar"
                elif cell_a == 42:
                    try:
                        if abs(valor_coluna_h - ipi_a_recup) <= 0.10:
                            row[8].value = "OK"
                        else:
                            row[8].value = "Verificar"
                    except NameError:
                        print("Erro: IPI_a_recup não definido. Continuando a execução...")
                        row[8].value = "Verificar"
                elif cell_a == 46:
                    try:
                        if abs(valor_coluna_h - cofins_recup) <= 0.10:
                            row[8].value = "OK"
                        else:
                            row[8].value = "Verificar"
                    except NameError:
                        print("Erro: cofins_recup não definido. Continuando a execução...")
                        row[8].value = "Verificar"
            except TypeError:
                continue

def perform_actions():
    company_code, month_year, company_name = get_user_input()
    html_path_ipi_a_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração IPI Sequência 22 - Ordem 2.htm"
    ipi_a_recup = extract_saldo_credor(html_path_ipi_a_recup)
    print(f"IPI a recup: {ipi_a_recup:.2f}".replace('.', ','))

    html_path_cofins_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração Cofins Sequência 22 - Ordem 3.htm"
    cofins_recup = extract_saldo_credor_cofins_recup(html_path_cofins_recup)
    print(f"Cofins a recup: {cofins_recup:.2f}".replace('.', ','))

    excel_path = f'C:\\projeto\\planilhas\\balancete\\CONCILIACAO_{company_code}_{month_year}.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    numeros_procurados = [41, 42, 46]
    compare_and_write_to_excel(ws, numeros_procurados, None, ipi_a_recup, cofins_recup)

    # Salvar as alterações de volta no arquivo Excel
    wb.save(excel_path)

def perform_actions_cofins_a_recup():
    company_code, month_year, company_name = get_user_input()
    html_path_cofins_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração Cofins Sequência 22 - Ordem 3.htm"
    cofins_recup = extract_saldo_credor_cofins_recup(html_path_cofins_recup)
    print(f"Cofins a recup: {cofins_recup:.2f}".replace('.', ','))

    excel_path = f'C:\\projeto\\planilhas\\balancete\\CONCILIACAO_{company_code}_{month_year}.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    numeros_procurados = [46]
    compare_and_write_to_excel(ws, numeros_procurados, None, None, cofins_recup)

    # Salvar as alterações de volta no arquivo Excel
    wb.save(excel_path)

